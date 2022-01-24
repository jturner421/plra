import os
import random
from decimal import Decimal

import openpyxl
import xlrd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font
from xlrd.timemachine import xrange


def _create_styles(workbook):
    data = NamedStyle(name='data')
    data.font = Font(name='Calibri', bold=False, size=11, color="000000")
    workbook.add_named_style(data)
    return workbook


def open_xls_file(filename):
    """
    Reads in WI state XLS file format and populates an Openpyxl workbook object
    :param filename: XLS file format
    :return: XLSX file
    """
    # TODO: Check for UnicodeDecodeError and handle
    book = xlrd.open_workbook(filename)
    book_sheet_names = book.sheet_names()
    index = 0
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.active

    # copy rows from xlrd to openpyxl sheet object
    for row in xrange(0, nrows):
        for col in xrange(0, ncols):
            sheet1.cell(row=row + 1, column=col + 1).value = sheet.cell_value(row, col)

    return book1


def format_case_num(case_num):
    """
    Identifies and formats case numbers for JIFMS lookup
    :param case_num: prisoner case number
    :return: CCAM formatted case number
    """
    # Check if multi-defendant case
    try:
        case_num_split = str.split(case_num, '-')
        if len(case_num_split) > 3:
            formatted_case_num = f"DWIW3{case_num_split[0]}{case_num_split[1]}{case_num_split[2].zfill(6)}-{case_num_split[3]}"
            return formatted_case_num
        else:
            formatted_case_num = f"DWIW3{case_num_split[0]}{case_num_split[1]}{case_num_split[2].zfill(6)}-001"
            return formatted_case_num
    except IndexError:
        pass

    except TypeError:
        print('No valid case found')
        pass


def create_output_file(check_date, check_num, output_path):
    """
    Creates blank Excel file used to populate payee date for later upload as CCAM batch job
    :param check_date: Date check received by the court
    :param check_num: state check number
    :param output_path: path for output file
    :return: Formatted Excel file
    """
    wb = openpyxl.Workbook()
    wb = _create_styles(wb)
    sheet = wb.active
    sheet.title = 'PLRA'
    sheet['A1'] = 'Control No.'
    sheet['B1'] = 'Agency Tracking ID.'
    sheet['C1'] = 'ACCOUNT HOLDER NAME (20 character max)'
    sheet['D1'] = 'EFFECTIVE DATE'
    sheet['E1'] = 'TRANSACTION AMOUNT'
    sheet['F1'] = 'DEPOSIT NO (ie CD102815)'
    sheet['G1'] = 'DOCKET / COURT NO.'
    sheet['H1'] = 'TOTAL OWED'
    sheet['I1'] = 'TOTAL COLLECTED'
    sheet['J1'] = 'TOTAL OUTSTANDING'
    sheet['K1'] = 'OVERPAYMENT'

    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 13
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 21
    sheet.column_dimensions['H'].width = 12
    sheet.column_dimensions['I'].width = 11
    sheet.column_dimensions['J'].width = 14
    sheet.column_dimensions['K'].width = 17

    for x in [chr(i) for i in range(ord('A'), ord('L'))]:
        col = f"{x}1"
        sheet[col].style = 'Accent4'
        sheet[col].font = Font(name='Calibri', bold=True, size=11, color="FFFFFF")

    check_date = str.split(check_date, '/')
    file = f"{check_date[2]}.{check_date[0]}.{check_date[1]}_Check_{check_num}_Upload.xlsx"
    file = f"{output_path}/{file}"
    if os.path.isfile(file):
        os.remove(file)
        wb.save(file)
    else:
        wb.save(file)
    return file


def create_output_path(file):
    # Construct  output path in same directory as input file
    tmp_output_file_path = str.split(file, '/')
    del tmp_output_file_path[-1]
    output_file_path = '/'.join(tmp_output_file_path)
    return output_file_path


def get_shortened_name(name):
    """
    Shorten name to 20 characters or less. Names are shortened as follows:
    Names with first, two middle and one last e.g. Romelo Bob Russel Booker are shortened to first name,
        1st character of second name and last name
    Names with first middle and last are shortened to first name, 1st character of second name and last name
    If still over 20 characters, drop the middle initial
    Hyphenated names ae shortened e.g. Helson Pabon-Gonzalez are shortened to first name, 1st initial to left of
        hyphenation, last name
    :param name: Name from check
    :return:Name of 20 characters or less
    """

    # Todo - Format names from Last, First M to First M Last as DOJ name format breaks this algorithm
    while len(name) > 20:
        split_name = str.split(name, ' ')
        if len(split_name) == 4:
            # Drop on the first of the two middle names
            revised_split_name = [split_name[0], split_name[1][0], split_name[2]]
            shortened_name = " ".join(revised_split_name)
        elif len(split_name) == 3:
            if len(split_name[1]) > 1:
                middle_name = split_name[1][0]
                revised_split_name = [split_name[0], middle_name, split_name[2]]
            else:  # Drop middle initial if still over 20 characters
                revised_split_name = [split_name[0], split_name[2]]
            shortened_name = " ".join(revised_split_name)
        else:
            # check for hyphenation
            try:
                split_name = str.split(name, ' ')
                # split the hyphenation
                hyph_lastname = str.split(split_name[1], '-')
                # shorten the left part of the hyphenated name
                hyph_initial_lastname = hyph_lastname[0][0]
                hyph_lastname = f"{hyph_initial_lastname}-{hyph_lastname[1]}"
                revised_split_name = [split_name[0], hyph_lastname]
                shortened_name = " ".join(revised_split_name)
            except:
                print(f"{name} cannot be shortened to comply with CCAM")
        name = shortened_name
    return name


def write_rows_to_output_file(file, payee_list, deposit_num, effective_date):
    """
    Populates pre-formatted Excel with payee information for upload to CCAM
    :param file: pre-formatted Excel file
    :param payee_list: list of payees from a specific check
    :param deposit_num: deposit number
    :param effective_date: date file created
    :return: None
    """
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name('PLRA')
    rownum = 2  # skip first row for header
    for p in payee_list:
        # Transaction has 2 dictionary keys: Prisoner and Case
        if len(p) ==2:
            sheet.cell = _transaction_row(deposit_num, effective_date, p, rownum, sheet)
        else:
            sheet.cell = _overpayment_row(deposit_num, effective_date, p, rownum, sheet)
        rownum += 1

    wb.save(file)
    print(f"The PLRA upload file has been saved as {file}\n")


def _transaction_row(deposit_num, effective_date, p, rownum, sheet):
    # Control numbers need to be randomized to ensure that a number is not duplicated if a payee is on multiple
    # checks for the same day
    sheet.cell(row=rownum, column=1).value = random.randrange(0, 999, 1)
    sheet.cell(row=rownum, column=2).value = int(p['prisoner'].doc_num)
    # Check length of name to fit within CCAM batch upload constraints
    try:
        if len(p['prisoner'].check_name) <= 20:
            sheet.cell(row=rownum, column=3).value = p['prisoner'].check_name
        else:
            shortened_name = get_shortened_name(p['prisoner'].check_name)
            sheet.cell(row=rownum, column=3).value = shortened_name
    except TypeError as error:
        print(f'{p.check_name} threw {error}')
    try:
        sheet.cell(row=rownum, column=4).value = effective_date
        sheet.cell(row=rownum, column=5).value = Decimal(p['case'].transaction.amount_paid)
        sheet.cell(row=rownum, column=6).value = deposit_num
        sheet.cell(row=rownum, column=7).value = str.upper(p['case'].ccam_case_num)
        sheet.cell(row=rownum, column=8).value = p['case'].balance.amount_assessed
        sheet.cell(row=rownum, column=9).value = p['case'].balance.amount_collected
        sheet.cell(row=rownum, column=10).value = p['case'].balance.amount_owed
    except AttributeError:
        sheet.cell(row=rownum, column=4).value = effective_date
        sheet.cell(row=rownum, column=5).value = Decimal(p['prisoner'].amount_paid)
        sheet.cell(row=rownum, column=6).value = deposit_num
        sheet.cell(row=rownum, column=7).value = p['case'].ecf_case_num.upper()
        sheet.cell(row=rownum, column=8).value = 0
        sheet.cell(row=rownum, column=9).value = 0
        sheet.cell(row=rownum, column=10).value = 0
        sheet.cell(row=rownum, column=11).value = -Decimal(p['prisoner'].amount_paid)
    for c in range(1, 9):
        if c not in [5, 8, 9, 10]:
            sheet.cell(row=rownum, column=c).style = 'data'
        else:
            sheet.cell(row=rownum, column=c).number_format = '$#,##0.00'
    return sheet.cell


def _overpayment_row(deposit_num, effective_date, p, rownum, sheet):
    # Control numbers need to be randomized to ensure that a number is not duplicated if a payee is on multiple
    # checks for the same day
    sheet.cell(row=rownum, column=1).value = random.randrange(0, 999, 1)
    sheet.cell(row=rownum, column=2).value = int(p['prisoner'].doc_num)
    sheet.cell(row=rownum, column=3).value = p['prisoner'].check_name
    sheet.cell(row=rownum, column=4).value = effective_date
    sheet.cell(row=rownum, column=5).value = Decimal(p['prisoner'].overpayment['transaction amount'])
    sheet.cell(row=rownum, column=6).value = deposit_num
    sheet.cell(row=rownum, column=7).value = p['prisoner'].overpayment['ccam_case_num']
    # sheet.cell(row=rownum, column=8).value = p['prisoner'].overpayment['assessed']
    # sheet.cell(row=rownum, column=9).value = p['prisoner'].overpayment['collected']
    # sheet.cell(row=rownum, column=9).value = p['prisoner'].overpayment['outstanding']
    sheet.cell(row=rownum, column=11).value = Decimal(p['prisoner'].refund)
    return sheet.cell
