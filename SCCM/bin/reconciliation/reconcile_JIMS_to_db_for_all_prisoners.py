"""
Performs reconciliation between JIFMS and application database

This module compares balances for all prisoners in the application database against JIFMS CCAM data and outputs the
results to a Microsoft Excel file for review and subsequent processing.  The output file serves as input to the
update_db_balances_from_JIFMS_for_all_prisoners module.
"""
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import datetime
from datetime import datetime, timedelta
import warnings

#import keyring
import openpyxl
import pandas as pd
import requests
from sqlalchemy import and_, func
from sqlalchemy.exc import SAWarning
from SCCM.bin import convert_to_excel as cte
from SCCM.config import config
from SCCM.data.case_balance import CaseBalance
from SCCM.data.court_cases import CourtCase
from SCCM.data.db_session import DbSession


def get_ccam_account_information(case_num, session, base_url):
    """
    API call to JIFMS to retrieve case balance information for a single case
    :param case_num: JIFMS formatted case number
    :param session: requests session object
    :param base_url: API url without search term
    :return: JSON object with case information
    """

    case_string_split = str.split(case_num, '-')
    case = f'{str.upper(case_string_split[0])}-{case_string_split[1]}'

    response = session.get(
        base_url,
        params={'caseNumberList': case},
        verify=False
    )
    return response.json()


def create_output_file(output_path):
    """
    Creates Excel file used to store reconciliation data

    :param output_path: path to create file
    :return: file storage location
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Compare PLRA Totals'
    sheet['A1'] = 'Prisoner Name'
    sheet['B1'] = 'DOC Number'
    sheet['C1'] = 'CaseBase ID'
    sheet['D1'] = 'CaseBase Number'
    sheet['E1'] = 'JIFMS Formatted CaseBase Number'
    sheet['F1'] = 'DB Amount Assessed'
    sheet['G1'] = 'DB Amount Collected'
    sheet['H1'] = 'DB Amount Owed'
    sheet['I1'] = 'API Amount Assessed'
    sheet['J1'] = 'API Amount Collected'
    sheet['k1'] = 'API Amount Owed'
    sheet['L1'] = 'Amount Assessed Diff'
    sheet['M1'] = 'Amount Collected Diff'
    sheet['N1'] = 'Amount Owed Diff'

    file_name = 'PLRA Compare.xlsx'
    output_path = f'{output_path}/{file_name}'
    wb.save(output_path)
    return output_path


def write_rows_to_output_file(file, comparison_list):
    """
    Writes balance information to Excel file
    :param file: path to Excel file
    :param comparison_list: reconciliation data row
    """

    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name('Compare PLRA Totals')
    rownum = sheet.max_row + 1  # skip first row for header
    sheet.cell(row=rownum, column=1).value = comparison_list[0]
    sheet.cell(row=rownum, column=2).value = comparison_list[1]
    sheet.cell(row=rownum, column=3).value = comparison_list[2]
    sheet.cell(row=rownum, column=4).value = comparison_list[3]
    sheet.cell(row=rownum, column=5).value = comparison_list[4]
    sheet.cell(row=rownum, column=6).value = comparison_list[5]
    sheet.cell(row=rownum, column=7).value = comparison_list[6]
    sheet.cell(row=rownum, column=8).value = comparison_list[7]
    sheet.cell(row=rownum, column=9).value = comparison_list[8]
    sheet.cell(row=rownum, column=10).value = comparison_list[9]
    sheet.cell(row=rownum, column=11).value = comparison_list[10]
    sheet.cell(row=rownum, column=12).value = comparison_list[11]
    sheet.cell(row=rownum, column=13).value = comparison_list[12]
    sheet.cell(row=rownum, column=14).value = comparison_list[13]

    wb.save(file)


def get_ccam_balances(payments):
    """
    Totals CCAM payment lines and updates prison object with payment information as a list
    :param payments: List of individual payment lines for a case
    :return: dataframe of case balance
    """
    # FIXME : Need to filter out OVP line (overpayment) so as not to overinflate amount assessed
    try:
        df = pd.DataFrame.from_dict(payments['data'])
        party_code = {'Party Code': df['prty_cd'][0]}
        df = df.drop(
            ['ecf_case_num', 'case_titl', 'acct_cd', 'prty_num', 'prty_nm', 'scty_org', 'debt_typ_lnum', 'debt_typ',
             'last_updated', 'prty_cd'], axis=1)
        col_dict = {'prnc_owed': 'Total Owed', 'prnc_clld': 'Total Collected',
                    'totl_ostg': 'Total Outstanding'}
        df.columns = [col_dict.get(x, x) for x in df.columns]

        df = df.sum()
        df1 = pd.Series(party_code)
        df = df.append(df1)
        ccam_account_balance = pd.Series.to_dict(df)
        return ccam_account_balance
    except KeyError:
        pass
    except ValueError:
        pass



def main():
    # config_path = Path.cwd()
    config_path = Path('/Users/jwt/PycharmProjects/plra/SCCM')
    # config_file = config_path.parent.parent / 'config' / 'config.ini'
    config_file = config_path / 'config' / 'config.ini'
    configuration = config.initialize_config(str(config_file))
    prod_vars = config.get_prod_vars(configuration, 'PROD')
    network_base_dir = prod_vars['NETWORK_BASE_DIR']
    prod_db_path = prod_vars['NETWORK_DB_BASE_DIR']
    db_file_name = prod_vars['DATABASE_SQLite']
    db_file = f'{prod_db_path}{db_file_name}'
    db_session = DbSession.global_init(db_file)
    ccam_username = prod_vars['CCAM_USERNAME']
    base_url = prod_vars['CCAM_API']
    # reconciliation_file_path = prod_vars['RECONCILIATION_FILE_PATH']

    reconciliation_file_path = '/Users/jwt/PycharmProjects/plra/SCCM/bin/reconciliation'
    ccam_password = keyring.get_password("WIWCCA", ccam_username)
    session = requests.Session()
    session.auth = (ccam_username, ccam_password)
    cert_path = prod_vars['CLIENT_CERT_PATH']
    session.verify = cert_path

    # get all prisoners
    s = db_session
    date = input('Please enter date of last check:  ')
    # TODO : validate input
    date = datetime.strptime(date, '%Y-%m-%d')
    as_of_date = f'{date} 00:00:00.000000'
    cutoff_date = date - timedelta(days=7)
    as_of_date = date.strftime("%Y-%m-%d")
    cutoff_date = cutoff_date.strftime("%Y-%m-%d")
    # query db for prisoners where the updated balance is between cutoff date and reconciliation date
    warnings.filterwarnings('ignore', r".*support Decimal objects natively", SAWarning, r'^sqlalchemy\.sql\.sqltypes$')
    changed_cases = s.query(CourtCase, CaseBalance).filter(CourtCase.id == CaseBalance.court_case_id).filter(
        CaseBalance.updated_date >= cutoff_date).all()
    reconciliation_file = create_output_file(reconciliation_file_path)

    for p in changed_cases:
        db_case = p.CourtCase.ecf_case_num
        db_amount_assessed = p.CaseBalance.amount_assessed
        db_amount_collected = p.CaseBalance.amount_collected
        db_amount_owed = p.CaseBalance.amount_owed
        try:
            judgement_name = p.CourtCase.prisoner.judgment_name
        except AttributeError as e:
            judgement_name = ''
        except ValueError:
            pass
        case_id = p.CourtCase.id

        # get values from API
        formatted_case_num = cte.format_case_num(db_case.upper())
        try:
            print(f'Retreiving CCAM account balances for {p.CourtCase.prisoner.legal_name} - {p.CourtCase.ecf_case_num} ')
        except AttributeError:
            pass
        api_values = get_ccam_account_information(formatted_case_num, session, base_url)
        try:
            ccam_payments = get_ccam_balances(api_values)
            cents = Decimal('0.01')
            db_amount_assessed = Decimal(db_amount_assessed).quantize(cents, ROUND_HALF_UP)
            db_amount_collected = Decimal(db_amount_collected).quantize(cents, ROUND_HALF_UP)
            db_amount_owed = Decimal(db_amount_owed).quantize(cents, ROUND_HALF_UP)
            amount_assessed_diff = db_amount_assessed - Decimal(float(ccam_payments['Total Owed']))
            amount_collected_diff = db_amount_collected - Decimal(float(ccam_payments['Total Collected']))
            amount_owed_diff = db_amount_owed - Decimal(float(ccam_payments['Total Outstanding']))
        except TypeError:
            pass

        if int(amount_owed_diff) == 0:
            pass
        else:
            try:
                compare_line = [judgement_name, p[0].prisoner_doc_num, p.CourtCase.id, db_case, formatted_case_num,
                                db_amount_assessed,
                                db_amount_collected, db_amount_owed,
                                Decimal(float(ccam_payments['Total Owed'])).quantize(cents, ROUND_HALF_UP),
                                Decimal(float(ccam_payments['Total Collected'])).quantize(cents, ROUND_HALF_UP),
                                Decimal(float(ccam_payments['Total Outstanding'])).quantize(cents, ROUND_HALF_UP),
                                amount_assessed_diff.quantize(cents, ROUND_HALF_UP),
                                amount_collected_diff.quantize(cents, ROUND_HALF_UP),
                                amount_owed_diff.quantize(cents, ROUND_HALF_UP)
                                ]
                write_rows_to_output_file(reconciliation_file, compare_line)
            except TypeError as error:
                compare_line = [judgement_name, case_id, db_case, formatted_case_num, '', '', '', '', '', '',
                                error.args[0]]
                try:
                    write_rows_to_output_file(reconciliation_file, compare_line)
                except IndexError:
                    pass


if __name__ == '__main__':
    main()
